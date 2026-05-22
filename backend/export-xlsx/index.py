"""
Генерация xlsx-файла с данными и встроенными графиками через openpyxl.
Принимает JSON с данными месторождения, возвращает base64-encoded xlsx.
"""
import json
import base64
import io

def handler(event: dict, context) -> dict:
    cors_headers = {
        'Access-Control-Allow-Origin': '*',
        'Access-Control-Allow-Methods': 'GET, POST, PUT, DELETE, OPTIONS',
        'Access-Control-Allow-Headers': 'Content-Type, X-User-Id, X-Auth-Token',
        'Content-Type': 'application/json'
    }

    if event.get('httpMethod') == 'OPTIONS':
        return {'statusCode': 200, 'headers': cors_headers, 'body': ''}

    import openpyxl
    from openpyxl.chart import BarChart, LineChart, Reference, Series
    from openpyxl.chart.series import SeriesLabel
    from openpyxl.chart.label import DataLabel
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, numbers
    from openpyxl.chart import AreaChart
    from openpyxl.chart.reference import Reference

    body = event.get('body', '')
    data = json.loads(body) if body else {}

    field_name = data.get('field_name', 'Месторождение')
    years = data.get('years', [2020, 2021, 2022, 2023, 2024])

    # Данные для графика 1 (изменения %)
    chart1_rows = data.get('chart1_rows', [])
    # Данные для графика 2 (динамика)
    chart2_data = data.get('chart2_data', {})

    wb = openpyxl.Workbook()

    # ======== Лист 1: Анализ изменений (горизонтальные бар-чарты) ========
    ws1 = wb.active
    ws1.title = "Анализ изменений"

    # Заголовок
    ws1['A1'] = f"Анализ изменений показателей — {field_name}"
    ws1['A1'].font = Font(bold=True, size=13)
    ws1.merge_cells('A1:M1')

    # Заголовки таблицы
    header_row = 3
    ws1.cell(row=header_row, column=1, value="Показатель").font = Font(bold=True)
    for i, yr in enumerate(years):
        col = 2 + i
        ws1.cell(row=header_row, column=col, value=str(yr)).font = Font(bold=True)

    # Данные
    indicators = [
        "Добыча нефти", "Добыча жидкости", "Обводненность (весовая)",
        "Дейст. добывающий фонд", "Дейст. нагнетательный фонд",
        "Закачка", "Приемистость", "Дебит нефти", "Дебит жидкости"
    ]

    default_values = [
        [-5, -22, 9, 2, 10],
        [-2, -20, -2, 4, 12],
        [1, 1, -4, 1, 1],
        [-6, 0, 7, -3, -1],
        [12, -15, 12, 11, 24],
        [6, -10, 0, 7, 10],
        [3, 12, 6, -3, -8],
        [3, -10, 20, -6, 2],
        [5, -8, 8, -3, 4],
    ]

    for r_idx, (ind, defaults) in enumerate(zip(indicators, default_values)):
        row = header_row + 1 + r_idx
        ws1.cell(row=row, column=1, value=ind)
        row_data = chart1_rows[r_idx] if r_idx < len(chart1_rows) else defaults
        for c_idx, val in enumerate(row_data[:len(years)]):
            ws1.cell(row=row, column=2 + c_idx, value=val)

    data_min_row = header_row + 1
    data_max_row = header_row + len(indicators)

    # График 1: горизонтальные бар-чарты по годам
    for yr_idx, yr in enumerate(years):
        chart = BarChart()
        chart.type = "bar"
        chart.barDir = "bar"
        chart.title = str(yr)
        chart.style = 2
        chart.grouping = "clustered"
        chart.width = 8
        chart.height = 12

        yr_col = 2 + yr_idx
        data_ref = Reference(ws1, min_col=yr_col, max_col=yr_col,
                             min_row=data_min_row, max_row=data_max_row)
        cats = Reference(ws1, min_col=1, min_row=data_min_row, max_row=data_max_row)
        chart.add_data(data_ref)
        chart.set_categories(cats)
        chart.series[0].title = SeriesLabel(v=str(yr))

        # Серые и синие цвета как в оригинале
        neg_fill = "808080"
        pos_fill = "4472C4"
        chart.series[0].graphicalProperties.solidFill = pos_fill

        anchor_col = get_column_letter(8 + yr_idx * 9)
        ws1.add_chart(chart, f"{anchor_col}3")

    # ======== Лист 2: Динамика показателей ========
    ws2 = wb.create_sheet("Динамика показателей")

    ws2['A1'] = f"Динамика показателей — {field_name}"
    ws2['A1'].font = Font(bold=True, size=13)
    ws2.merge_cells('A1:H1')

    years2 = chart2_data.get('years', [2022, 2023, 2024, 2025])
    ws2.cell(row=3, column=1, value="Год").font = Font(bold=True)
    ws2.cell(row=3, column=2, value="Закачка, тыс.м3").font = Font(bold=True)
    ws2.cell(row=3, column=3, value="Добыча жидкости, тыс.т").font = Font(bold=True)
    ws2.cell(row=3, column=4, value="Добыча нефти, тыс.т").font = Font(bold=True)
    ws2.cell(row=3, column=5, value="Фонд добыв. скв., ед").font = Font(bold=True)
    ws2.cell(row=3, column=6, value="Фонд нагн. скв., ед").font = Font(bold=True)
    ws2.cell(row=3, column=7, value="Компенсация тек., %").font = Font(bold=True)

    zakachka = chart2_data.get('zakachka', [550, 1050, 1600, 2000])
    liquid = chart2_data.get('liquid', [0, 0, 0, 2100])
    oil = chart2_data.get('oil', [500, 1000, 1550, 1700])
    fond_dob = chart2_data.get('fond_dob', [10, 20, 20, 20])
    fond_nag = chart2_data.get('fond_nag', [10, 28, 30, 55])
    compensation = chart2_data.get('compensation', [45, 90, 108, 108])

    for i, yr in enumerate(years2):
        row = 4 + i
        ws2.cell(row=row, column=1, value=yr)
        ws2.cell(row=row, column=2, value=zakachka[i] if i < len(zakachka) else 0)
        ws2.cell(row=row, column=3, value=liquid[i] if i < len(liquid) else 0)
        ws2.cell(row=row, column=4, value=oil[i] if i < len(oil) else 0)
        ws2.cell(row=row, column=5, value=fond_dob[i] if i < len(fond_dob) else 0)
        ws2.cell(row=row, column=6, value=fond_nag[i] if i < len(fond_nag) else 0)
        ws2.cell(row=row, column=7, value=compensation[i] if i < len(compensation) else 0)

    d2_min = 4
    d2_max = 3 + len(years2)

    # AreaChart для закачки и жидкости
    area_chart = AreaChart()
    area_chart.title = "Динамика показателей"
    area_chart.style = 2
    area_chart.width = 20
    area_chart.height = 14
    area_chart.grouping = "stacked"

    cats2 = Reference(ws2, min_col=1, min_row=d2_min, max_row=d2_max)

    zakachka_ref = Reference(ws2, min_col=2, min_row=3, max_row=d2_max)
    liquid_ref = Reference(ws2, min_col=3, min_row=3, max_row=d2_max)
    oil_ref = Reference(ws2, min_col=4, min_row=3, max_row=d2_max)

    area_chart.add_data(zakachka_ref, titles_from_data=True)
    area_chart.add_data(liquid_ref, titles_from_data=True)
    area_chart.add_data(oil_ref, titles_from_data=True)
    area_chart.set_categories(cats2)

    area_chart.series[0].graphicalProperties.solidFill = "FFFF00"
    area_chart.series[1].graphicalProperties.solidFill = "ADD8E6"
    area_chart.series[2].graphicalProperties.solidFill = "C0504D"

    # BarChart для фондов
    bar_chart2 = BarChart()
    bar_chart2.type = "col"
    bar_chart2.style = 2
    bar_chart2.grouping = "clustered"
    bar_chart2.width = 20
    bar_chart2.height = 14

    fd_ref = Reference(ws2, min_col=5, min_row=3, max_row=d2_max)
    fn_ref = Reference(ws2, min_col=6, min_row=3, max_row=d2_max)
    bar_chart2.add_data(fd_ref, titles_from_data=True)
    bar_chart2.add_data(fn_ref, titles_from_data=True)
    bar_chart2.set_categories(cats2)
    bar_chart2.series[0].graphicalProperties.solidFill = "8B4513"
    bar_chart2.series[1].graphicalProperties.solidFill = "00BFFF"

    area_chart += bar_chart2

    # LineChart для компенсации (вторая ось Y)
    line_chart = LineChart()
    comp_ref = Reference(ws2, min_col=7, min_row=3, max_row=d2_max)
    line_chart.add_data(comp_ref, titles_from_data=True)
    line_chart.set_categories(cats2)
    line_chart.series[0].graphicalProperties.line.solidFill = "00008B"
    line_chart.series[0].graphicalProperties.line.width = 20000
    line_chart.y_axis.axId = 200
    line_chart.y_axis.crosses = "max"

    area_chart += line_chart
    ws2.add_chart(area_chart, "A10")

    # Сохранить в байты
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    xlsx_b64 = base64.b64encode(output.read()).decode('utf-8')

    return {
        'statusCode': 200,
        'headers': cors_headers,
        'body': json.dumps({'file': xlsx_b64, 'filename': f'{field_name}_report.xlsx'})
    }
