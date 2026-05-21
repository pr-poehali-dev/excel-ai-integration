"""
Генерирует xlsx-файл с данными + встроенной диаграммой через openpyxl.
Принимает: sheets (все листы файла), chart_sheet (имя листа с данными для графика), chart_type.
Возвращает base64-encoded xlsx.
"""

import json
import base64
import io
from openpyxl import Workbook
from openpyxl.chart import PieChart, BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


CHART_COLORS = [
    "34D399", "FBBF24", "60A5FA", "A78BFA",
    "F87171", "2DD4BF", "FB923C", "818CF8",
]

THIN = Side(style="thin", color="2D3748")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header_cell(cell):
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", fgColor="1A202C")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def style_data_cell(cell, row_idx: int):
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = BORDER
    if row_idx % 2 == 0:
        cell.fill = PatternFill("solid", fgColor="1E2A3A")


def handler(event: dict, context) -> dict:
    """Генерация Excel-файла со встроенным графиком"""

    if event.get("httpMethod") == "OPTIONS":
        return {
            "statusCode": 200,
            "headers": {
                "Access-Control-Allow-Origin": "*",
                "Access-Control-Allow-Methods": "POST, OPTIONS",
                "Access-Control-Allow-Headers": "Content-Type",
                "Access-Control-Max-Age": "86400",
            },
            "body": "",
        }

    body = json.loads(event.get("body") or "{}")
    sheets_data = body.get("sheets", [])       # [{name, data: [[...],...]}]
    chart_sheet_name = body.get("chart_sheet") # имя листа с данными для графика
    chart_type = body.get("chart_type", "pie") # pie | bar | line
    chart_title = body.get("chart_title", "График")

    if not sheets_data:
        return {
            "statusCode": 400,
            "headers": {"Access-Control-Allow-Origin": "*"},
            "body": json.dumps({"error": "Нет данных"}, ensure_ascii=False),
        }

    wb = Workbook()
    wb.remove(wb.active)  # удаляем дефолтный пустой лист

    chart_ws = None

    for sh in sheets_data:
        ws = wb.create_sheet(title=str(sh["name"])[:31])
        rows = sh.get("data", [])

        for ri, row in enumerate(rows):
            for ci, val in enumerate(row):
                if val is None:
                    continue
                cell = ws.cell(row=ri + 1, column=ci + 1, value=val)
                if ri == 0:
                    style_header_cell(cell)
                else:
                    style_data_cell(cell, ri)

        # Авто-ширина колонок
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

        ws.row_dimensions[1].height = 20

        if sh["name"] == chart_sheet_name:
            chart_ws = ws

    # Строим диаграмму на нужном листе
    if chart_ws is not None:
        data_rows = chart_ws.max_row
        data_cols = chart_ws.max_column

        if data_rows >= 2 and data_cols >= 2:
            # Метки (первая колонка, строки 2..N)
            labels = Reference(chart_ws, min_col=1, min_row=2, max_row=data_rows)
            # Значения (вторая колонка, строки 2..N)
            values = Reference(chart_ws, min_col=2, min_row=1, max_row=data_rows)

            if chart_type == "pie":
                chart = PieChart()
                chart.title = chart_title
                chart.style = 10
                chart.add_data(values, titles_from_data=True)
                chart.set_categories(labels)
                chart.dataLabels = None

                # Раскрашиваем сегменты
                series = chart.series[0]
                for idx in range(data_rows - 1):
                    pt = DataPoint(idx=idx)
                    color = CHART_COLORS[idx % len(CHART_COLORS)]
                    pt.graphicalProperties.solidFill = color
                    series.dPt.append(pt)

            elif chart_type == "bar":
                chart = BarChart()
                chart.type = "col"
                chart.title = chart_title
                chart.style = 10
                chart.add_data(values, titles_from_data=True)
                chart.set_categories(labels)

            else:
                chart = LineChart()
                chart.title = chart_title
                chart.style = 10
                chart.add_data(values, titles_from_data=True)
                chart.set_categories(labels)

            chart.width = 18
            chart.height = 12

            # Добавляем график на тот же лист, правее данных
            anchor_col = get_column_letter(data_cols + 2)
            chart_ws.add_chart(chart, f"{anchor_col}2")

    # Сохраняем в bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    b64 = base64.b64encode(buf.read()).decode()

    return {
        "statusCode": 200,
        "headers": {
            "Access-Control-Allow-Origin": "*",
            "Content-Type": "application/json",
        },
        "body": json.dumps({"xlsx_b64": b64}, ensure_ascii=False),
    }
